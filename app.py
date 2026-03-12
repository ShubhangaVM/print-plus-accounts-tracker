from flask import Flask, render_template, request, redirect, url_for, flash
from database import init_db, get_connection
from datetime import datetime, date

app = Flask(__name__)
app.secret_key = "printplus_secret_key"

init_db()

# ----------------------------
# Helpers
# ----------------------------
def today_str():
    return datetime.today().strftime("%Y-%m-%d")

def days_overdue(sale_date_str):
    """
    sale_date_str format: YYYY-MM-DD
    """
    try:
        d = datetime.strptime(sale_date_str, "%Y-%m-%d").date()
        return (date.today() - d).days
    except Exception:
        return 0

# ----------------------------
# Home / Dashboard
# ----------------------------
@app.route("/")
def home():
    conn = get_connection()
    cur = conn.cursor()
    try:
        # Total pending amount and count
        cur.execute("""
            SELECT COUNT(*) as cnt, SUM(amount - paid_amount) as total
            FROM sales
            WHERE (amount - paid_amount) > 0.001
        """)
        pending_row = cur.fetchone()

        # Today's sales total
        cur.execute("""
            SELECT COALESCE(SUM(amount), 0) as total
            FROM sales
            WHERE sale_date = ?
        """, (today_str(),))
        today_row = cur.fetchone()

    finally:
        conn.close()

    return render_template("home.html",
        pending_count=pending_row["cnt"] or 0,
        pending_total=float(pending_row["total"] or 0),
        today_total=float(today_row["total"] or 0),
    )

# ----------------------------
# Customers
# ----------------------------
@app.route("/customers", methods=["GET", "POST"])
def customers():
    conn = get_connection()
    cur = conn.cursor()

    try:
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            phone = (request.form.get("phone") or "").strip()
            address = (request.form.get("address") or "").strip()
            gst_no = (request.form.get("gst_no") or "").strip()

            if not name:
                flash("Customer name is required.", "error")
                return redirect(url_for("customers"))

            cur.execute(
                "INSERT INTO customers (name, phone, address, gst_no) VALUES (?, ?, ?, ?)",
                (name, phone, address, gst_no)
            )
            conn.commit()
            flash("Customer added ✅", "success")

        cur.execute("SELECT * FROM customers ORDER BY name ASC")
        rows = cur.fetchall()
    finally:
        conn.close()

    return render_template("customers.html", rows=rows)

@app.route("/customers/edit/<int:cid>", methods=["GET", "POST"])
def edit_customer(cid):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT * FROM customers WHERE id=?", (cid,))
        c = cur.fetchone()

        if not c:
            flash("Customer not found.", "error")
            return redirect(url_for("customers"))

        if request.method == "POST":
            name = request.form.get("name", "").strip()
            phone = request.form.get("phone", "").strip()
            address = request.form.get("address", "").strip()
            gst_no = request.form.get("gst_no", "").strip()

            cur.execute("""
                UPDATE customers
                SET name=?, phone=?, address=?, gst_no=?
                WHERE id=?
            """, (name, phone, address, gst_no, cid))

            conn.commit()
            flash("Customer updated ✅", "success")
            return redirect(url_for("customers"))
    finally:
        conn.close()

    return render_template("customer_edit.html", c=c)

@app.route("/customers/delete/<int:customer_id>", methods=["POST"])
def customer_delete(customer_id):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT COUNT(*) as cnt FROM sales WHERE customer_id = ?", (customer_id,))
        cnt = cur.fetchone()["cnt"]

        if cnt > 0:
            flash("Cannot delete. This customer has sales records.", "error")
            return redirect(url_for("customers"))

        cur.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
        conn.commit()
    finally:
        conn.close()

    flash("Customer deleted ✅", "success")
    return redirect(url_for("customers"))

# ----------------------------
# Sellers
# ----------------------------
@app.route("/sellers", methods=["GET", "POST"])
def sellers():
    conn = get_connection()
    cur = conn.cursor()

    try:
        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            phone = (request.form.get("phone") or "").strip()
            address = (request.form.get("address") or "").strip()
            gst_no = (request.form.get("gst_no") or "").strip()

            if not name:
                flash("Seller name is required.", "error")
                return redirect(url_for("sellers"))

            cur.execute(
                "INSERT INTO sellers (name, phone, address, gst_no) VALUES (?, ?, ?, ?)",
                (name, phone, address, gst_no)
            )
            conn.commit()
            flash("Seller added ✅", "success")

        cur.execute("SELECT * FROM sellers ORDER BY name ASC")
        rows = cur.fetchall()
    finally:
        conn.close()

    return render_template("sellers.html", rows=rows)

@app.route("/sellers/edit/<int:sid>", methods=["GET", "POST"])
def edit_seller(sid):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT * FROM sellers WHERE id=?", (sid,))
        s = cur.fetchone()

        if not s:
            flash("Seller not found.", "error")
            return redirect(url_for("sellers"))

        if request.method == "POST":
            name = request.form.get("name", "").strip()
            phone = request.form.get("phone", "").strip()
            address = request.form.get("address", "").strip()
            gst_no = request.form.get("gst_no", "").strip()

            cur.execute("""
                UPDATE sellers
                SET name=?, phone=?, address=?, gst_no=?
                WHERE id=?
            """, (name, phone, address, gst_no, sid))

            conn.commit()
            flash("Seller updated ✅", "success")
            return redirect(url_for("sellers"))
    finally:
        conn.close()

    return render_template("seller_edit.html", s=s)

@app.route("/sellers/delete/<int:seller_id>", methods=["POST"])
def seller_delete(seller_id):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT COUNT(*) as cnt FROM purchases WHERE seller_id = ?", (seller_id,))
        cnt = cur.fetchone()["cnt"]

        if cnt > 0:
            flash("Cannot delete. This seller has purchase records.", "error")
            return redirect(url_for("sellers"))

        cur.execute("DELETE FROM sellers WHERE id = ?", (seller_id,))
        conn.commit()
    finally:
        conn.close()

    flash("Seller deleted ✅", "success")
    return redirect(url_for("sellers"))

# ----------------------------
# Sales - Add
# ----------------------------
@app.route("/sales/add", methods=["GET", "POST"])
def sales_add():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM customers ORDER BY name ASC")
    customers = cur.fetchall()
    conn.close()

    if request.method == "POST":
        sale_date = request.form.get("sale_date") or today_str()
        # FIXED: was `customer_id = customer_id.strip()` which crashed — never read from form
        customer_id = (request.form.get("customer_id") or "").strip()
        bill_no = (request.form.get("bill_no") or "").strip()
        amount = request.form.get("amount") or 0
        paid_amount = request.form.get("paid_amount") or 0
        notes = (request.form.get("notes") or "").strip()

        if not customer_id:
            flash("Please select a customer.", "error")
            return redirect(url_for("sales_add"))

        if not bill_no:
            flash("Bill number is required.", "error")
            return redirect(url_for("sales_add"))

        try:
            amount = float(amount)
            paid_amount = float(paid_amount)
        except ValueError:
            flash("Amount fields must be numbers.", "error")
            return redirect(url_for("sales_add"))

        if paid_amount > amount:
            flash("Paid amount cannot be more than total amount.", "error")
            return redirect(url_for("sales_add"))

        conn = get_connection()
        cur = conn.cursor()
        try:
            cur.execute("""
            INSERT INTO sales (sale_date, customer_id, bill_no, amount, paid_amount, notes)
            VALUES (?, ?, ?, ?, ?, ?)
            """, (sale_date, customer_id, bill_no, amount, paid_amount, notes))
            conn.commit()
        finally:
            conn.close()

        flash("Sale entry saved ✅", "success")
        return redirect(url_for("sales_pending"))

    return render_template("sales_add.html", today=today_str(), customers=customers)

# ----------------------------
# Sales - Pending
# ----------------------------
@app.route("/sales/pending")
def sales_pending():
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
        SELECT sales.*, customers.name AS customer_name
        FROM sales
        JOIN customers ON sales.customer_id = customers.id
        ORDER BY sale_date DESC, sales.id DESC
        """)
        rows = cur.fetchall()
    finally:
        conn.close()

    pending = []
    for r in rows:
        amount = float(r["amount"] or 0)
        paid = float(r["paid_amount"] or 0)
        pending_amt = amount - paid
        if pending_amt > 0.001:
            pending.append({
                "id": r["id"],
                "sale_date": r["sale_date"],
                "customer_name": r["customer_name"],
                "bill_no": r["bill_no"],
                "amount": amount,
                "paid": paid,
                "pending": pending_amt,
                "due_days": days_overdue(r["sale_date"]),
                "notes": r["notes"] or ""
            })

    total_pending = sum([x["pending"] for x in pending])
    pending.sort(key=lambda x: x["due_days"], reverse=True)

    return render_template("sales_pending.html", rows=pending, total_pending=total_pending)

# ----------------------------
# Sales - All
# ----------------------------
@app.route("/sales")
def sales_list():
    q = request.args.get("q", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    try:
        if q:
            cur.execute("""
                SELECT sales.*, customers.name AS customer_name
                FROM sales
                JOIN customers ON sales.customer_id = customers.id
                WHERE customers.name LIKE ?
                   OR sales.bill_no LIKE ?
                   OR sales.notes LIKE ?
                ORDER BY sale_date DESC, sales.id DESC
            """, (f"%{q}%", f"%{q}%", f"%{q}%"))
        else:
            cur.execute("""
                SELECT sales.*, customers.name AS customer_name
                FROM sales
                JOIN customers ON sales.customer_id = customers.id
                ORDER BY sale_date DESC, sales.id DESC
            """)
        rows = cur.fetchall()
    finally:
        conn.close()

    data = []
    for r in rows:
        amount = float(r["amount"] or 0)
        paid = float(r["paid_amount"] or 0)
        pending_amt = amount - paid
        data.append({
            "id": r["id"],
            "sale_date": r["sale_date"],
            "customer_name": r["customer_name"],
            "bill_no": r["bill_no"],
            "amount": amount,
            "paid": paid,
            "pending": pending_amt,
            "due_days": days_overdue(r["sale_date"]),
            "notes": r["notes"] or ""
        })

    return render_template("sales_list.html", rows=data, q=q)

# ----------------------------
# Sales - Edit payment (with payment history log)
# ----------------------------
@app.route("/sales/payment/<int:sale_id>", methods=["GET", "POST"])
def sales_edit_payment(sale_id):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
        SELECT sales.*, customers.name AS customer_name
        FROM sales
        JOIN customers ON sales.customer_id = customers.id
        WHERE sales.id = ?
        """, (sale_id,))
        row = cur.fetchone()

        if not row:
            flash("Sale not found.", "error")
            return redirect(url_for("sales_pending"))

        if request.method == "POST":
            new_payment = request.form.get("new_payment") or 0
            payment_notes = (request.form.get("payment_notes") or "").strip()

            try:
                new_payment = float(new_payment)
            except ValueError:
                flash("Payment amount must be a number.", "error")
                return redirect(url_for("sales_edit_payment", sale_id=sale_id))

            if new_payment <= 0:
                flash("Payment amount must be greater than zero.", "error")
                return redirect(url_for("sales_edit_payment", sale_id=sale_id))

            amount = float(row["amount"] or 0)
            current_paid = float(row["paid_amount"] or 0)

            if current_paid + new_payment > amount:
                flash(f"Payment exceeds remaining balance of ₹{amount - current_paid:.2f}.", "error")
                return redirect(url_for("sales_edit_payment", sale_id=sale_id))

            # Log the payment in the payments table
            cur.execute("""
                INSERT INTO payments (sale_id, payment_date, amount, notes)
                VALUES (?, ?, ?, ?)
            """, (sale_id, today_str(), new_payment, payment_notes))

            # Update the running total on the sale
            cur.execute("UPDATE sales SET paid_amount = paid_amount + ? WHERE id = ?", (new_payment, sale_id))
            conn.commit()
            flash("Payment recorded ✅", "success")
            return redirect(url_for("sales_pending"))

        # Load payment history for this sale
        cur.execute("""
            SELECT * FROM payments WHERE sale_id = ? ORDER BY payment_date ASC, id ASC
        """, (sale_id,))
        payment_history = cur.fetchall()

    finally:
        conn.close()

    amount = float(row["amount"] or 0)
    paid = float(row["paid_amount"] or 0)

    return render_template(
        "sales_edit_payment.html",
        row=row,
        amount=amount,
        paid=paid,
        pending=amount - paid,
        due_days=days_overdue(row["sale_date"]),
        payment_history=payment_history
    )

# ----------------------------
# Sales - Mark Fully Paid (one-click shortcut)
# ----------------------------
@app.route("/sales/payment/<int:sale_id>/markpaid", methods=["POST"])
def sales_mark_paid(sale_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT amount, paid_amount FROM sales WHERE id = ?", (sale_id,))
        row = cur.fetchone()
        if not row:
            flash("Sale not found.", "error")
            return redirect(url_for("sales_pending"))

        remaining = float(row["amount"]) - float(row["paid_amount"])
        if remaining > 0:
            # Log the final payment
            cur.execute("""
                INSERT INTO payments (sale_id, payment_date, amount, notes)
                VALUES (?, ?, ?, ?)
            """, (sale_id, today_str(), remaining, "Marked as fully paid"))
            cur.execute("UPDATE sales SET paid_amount = amount WHERE id = ?", (sale_id,))
            conn.commit()
    finally:
        conn.close()
    flash("Bill marked as fully paid ✅", "success")
    return redirect(url_for("sales_pending"))

# ----------------------------
# Payment history - delete a single payment entry
# ----------------------------
@app.route("/sales/payment/entry/delete/<int:payment_id>", methods=["POST"])
def payment_entry_delete(payment_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM payments WHERE id = ?", (payment_id,))
        p = cur.fetchone()
        if not p:
            flash("Payment entry not found.", "error")
            return redirect(url_for("sales_pending"))

        sale_id = p["sale_id"]
        amount = float(p["amount"])

        # Remove payment log entry and reduce paid_amount on the sale
        cur.execute("DELETE FROM payments WHERE id = ?", (payment_id,))
        cur.execute("UPDATE sales SET paid_amount = MAX(0, paid_amount - ?) WHERE id = ?", (amount, sale_id))
        conn.commit()
        flash("Payment entry removed ✅", "success")
        return redirect(url_for("sales_edit_payment", sale_id=sale_id))
    finally:
        conn.close()

# ----------------------------
# Reports
# ----------------------------
@app.route("/reports")
def reports():
    from datetime import datetime

    selected_year = request.args.get("year", datetime.today().strftime("%Y"))
    selected_month = request.args.get("month", "")  # empty = full year

    conn = get_connection()
    cur = conn.cursor()

    try:
        # Build date filter
        if selected_month:
            date_prefix = f"{selected_year}-{selected_month.zfill(2)}"
            sales_filter = "WHERE strftime('%Y-%m', sale_date) = ?"
            purchases_filter = "WHERE strftime('%Y-%m', purchase_date) = ?"
            filter_val = date_prefix
        else:
            sales_filter = "WHERE strftime('%Y', sale_date) = ?"
            purchases_filter = "WHERE strftime('%Y', purchase_date) = ?"
            filter_val = selected_year

        # Summary stats
        cur.execute(f"SELECT COUNT(*) as cnt, SUM(amount) as total, SUM(paid_amount) as paid FROM sales {sales_filter}", (filter_val,))
        sales_summary = cur.fetchone()

        cur.execute(f"SELECT COUNT(*) as cnt, SUM(amount) as total FROM purchases {purchases_filter}", (filter_val,))
        purchases_summary = cur.fetchone()

        # Per-customer breakdown
        cur.execute(f"""
            SELECT customers.name, COUNT(*) as bills, SUM(sales.amount) as total,
                   SUM(sales.paid_amount) as paid, SUM(sales.amount - sales.paid_amount) as pending
            FROM sales
            JOIN customers ON sales.customer_id = customers.id
            {sales_filter}
            GROUP BY customers.id
            ORDER BY total DESC
        """, (filter_val,))
        customer_breakdown = cur.fetchall()

        # Available years for filter dropdown
        cur.execute("SELECT DISTINCT strftime('%Y', sale_date) as yr FROM sales ORDER BY yr DESC")
        years = [r["yr"] for r in cur.fetchall() if r["yr"]]
        if selected_year not in years:
            years.insert(0, selected_year)

    finally:
        conn.close()

    total_sales = float(sales_summary["total"] or 0)
    total_paid = float(sales_summary["paid"] or 0)
    total_pending = total_sales - total_paid
    total_purchases = float(purchases_summary["total"] or 0)
    profit = total_sales - total_purchases

    months = [
        ("01","January"),("02","February"),("03","March"),("04","April"),
        ("05","May"),("06","June"),("07","July"),("08","August"),
        ("09","September"),("10","October"),("11","November"),("12","December")
    ]

    return render_template("reports.html",
        selected_year=selected_year,
        selected_month=selected_month,
        years=years,
        months=months,
        sales_count=sales_summary["cnt"] or 0,
        total_sales=total_sales,
        total_paid=total_paid,
        total_pending=total_pending,
        purchases_count=purchases_summary["cnt"] or 0,
        total_purchases=total_purchases,
        profit=profit,
        customer_breakdown=customer_breakdown
    )

# ----------------------------
# PDF Invoice
# ----------------------------
@app.route("/sales/invoice/<int:sale_id>")
def sales_invoice(sale_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT sales.*, customers.name AS customer_name,
                   customers.phone AS customer_phone,
                   customers.address AS customer_address,
                   customers.gst_no AS customer_gst
            FROM sales
            JOIN customers ON sales.customer_id = customers.id
            WHERE sales.id = ?
        """, (sale_id,))
        sale = cur.fetchone()

        if not sale:
            flash("Sale not found.", "error")
            return redirect(url_for("sales_list"))

        cur.execute("SELECT * FROM payments WHERE sale_id = ? ORDER BY payment_date ASC", (sale_id,))
        payment_history = cur.fetchall()
    finally:
        conn.close()

    amount = float(sale["amount"] or 0)
    paid = float(sale["paid_amount"] or 0)

    return render_template("invoice.html",
        sale=sale,
        amount=amount,
        paid=paid,
        pending=amount - paid,
        payment_history=payment_history
    )

# ----------------------------
# Sales - Delete
# ----------------------------
@app.route("/sales/delete/<int:sale_id>", methods=["POST"])
def sales_delete(sale_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM sales WHERE id = ?", (sale_id,))
        conn.commit()
    finally:
        conn.close()
    flash("Sale deleted ✅", "success")
    return redirect(url_for("sales_list"))

# ----------------------------
# Purchases - Add
# ----------------------------
@app.route("/purchases/add", methods=["GET", "POST"])
def purchases_add():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM sellers ORDER BY name ASC")
    sellers = cur.fetchall()
    conn.close()

    if request.method == "POST":
        purchase_date = request.form.get("purchase_date") or today_str()
        seller_id = request.form.get("seller_id")
        payment_method = request.form.get("payment_method") or ""
        payment_ref = (request.form.get("payment_ref") or "").strip()
        bill_no = (request.form.get("bill_no") or "").strip()
        product_description = (request.form.get("product_description") or "").strip()
        amount = request.form.get("amount") or 0
        notes = (request.form.get("notes") or "").strip()

        if not seller_id:
            flash("Please select a seller.", "error")
            return redirect(url_for("purchases_add"))

        try:
            amount = float(amount)
        except ValueError:
            flash("Amount must be a number.", "error")
            return redirect(url_for("purchases_add"))

        conn = get_connection()
        cur = conn.cursor()
        try:
            cur.execute("""
            INSERT INTO purchases
            (purchase_date, seller_id, payment_method, payment_ref, bill_no, product_description, amount, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (purchase_date, seller_id, payment_method, payment_ref, bill_no, product_description, amount, notes))
            conn.commit()
        finally:
            conn.close()

        flash("Purchase saved ✅", "success")
        return redirect(url_for("purchases_list"))

    return render_template("purchases_add.html", today=today_str(), sellers=sellers)

# ----------------------------
# Purchases - List
# ----------------------------
@app.route("/purchases")
def purchases_list():
    q = request.args.get("q", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    try:
        if q:
            cur.execute("""
                SELECT purchases.*, sellers.name AS seller_name
                FROM purchases
                JOIN sellers ON purchases.seller_id = sellers.id
                WHERE sellers.name LIKE ?
                   OR purchases.bill_no LIKE ?
                   OR purchases.product_description LIKE ?
                   OR purchases.payment_ref LIKE ?
                ORDER BY purchases.purchase_date DESC, purchases.id DESC
            """, (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"))
        else:
            cur.execute("""
                SELECT purchases.*, sellers.name AS seller_name
                FROM purchases
                JOIN sellers ON purchases.seller_id = sellers.id
                ORDER BY purchases.purchase_date DESC, purchases.id DESC
            """)
        rows = cur.fetchall()
    finally:
        conn.close()

    return render_template("purchases_list.html", rows=rows, q=q)

# ----------------------------
# Purchases - Delete
# ----------------------------
@app.route("/purchases/delete/<int:purchase_id>", methods=["POST"])
def purchases_delete(purchase_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM purchases WHERE id = ?", (purchase_id,))
        conn.commit()
    finally:
        conn.close()
    flash("Purchase deleted ✅", "success")
    return redirect(url_for("purchases_list"))

# ----------------------------
# Purchases - Edit
# ----------------------------
@app.route("/purchases/edit/<int:purchase_id>", methods=["GET", "POST"])
def purchases_edit(purchase_id):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM sellers ORDER BY name ASC")
    sellers = cur.fetchall()

    try:
        cur.execute("SELECT * FROM purchases WHERE id = ?", (purchase_id,))
        p = cur.fetchone()

        if not p:
            flash("Purchase not found.", "error")
            return redirect(url_for("purchases_list"))

        if request.method == "POST":
            purchase_date = request.form.get("purchase_date") or today_str()
            seller_id = request.form.get("seller_id")
            payment_method = request.form.get("payment_method") or ""
            payment_ref = (request.form.get("payment_ref") or "").strip()
            bill_no = (request.form.get("bill_no") or "").strip()
            product_description = (request.form.get("product_description") or "").strip()
            amount = request.form.get("amount") or 0
            notes = (request.form.get("notes") or "").strip()

            if not seller_id:
                flash("Please select a seller.", "error")
                return redirect(url_for("purchases_edit", purchase_id=purchase_id))

            try:
                amount = float(amount)
            except ValueError:
                flash("Amount must be a number.", "error")
                return redirect(url_for("purchases_edit", purchase_id=purchase_id))

            cur.execute("""
                UPDATE purchases
                SET purchase_date=?, seller_id=?, payment_method=?, payment_ref=?,
                    bill_no=?, product_description=?, amount=?, notes=?
                WHERE id=?
            """, (purchase_date, seller_id, payment_method, payment_ref,
                    bill_no, product_description, amount, notes, purchase_id))
            conn.commit()
            flash("Purchase updated ✅", "success")
            return redirect(url_for("purchases_list"))

    finally:
        conn.close()

    return render_template("purchases_edit.html", p=p, sellers=sellers, today=today_str())


# ----------------------------
# Excel Export - Reports
# ----------------------------
@app.route("/reports/export")
def reports_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    selected_year  = request.args.get("year",  datetime.today().strftime("%Y"))
    selected_month = request.args.get("month", "")

    conn = get_connection()
    cur  = conn.cursor()
    try:
        if selected_month:
            filter_val       = f"{selected_year}-{selected_month.zfill(2)}"
            sales_filter     = "WHERE strftime('%Y-%m', sale_date) = ?"
            purchases_filter = "WHERE strftime('%Y-%m', purchase_date) = ?"
            period_label     = f"{selected_year}-{selected_month.zfill(2)}"
        else:
            filter_val       = selected_year
            sales_filter     = "WHERE strftime('%Y', sale_date) = ?"
            purchases_filter = "WHERE strftime('%Y', purchase_date) = ?"
            period_label     = selected_year

        cur.execute(f"SELECT COUNT(*) as cnt, SUM(amount) as total, SUM(paid_amount) as paid FROM sales {sales_filter}", (filter_val,))
        sales_summary = cur.fetchone()

        cur.execute(f"SELECT COUNT(*) as cnt, SUM(amount) as total FROM purchases {purchases_filter}", (filter_val,))
        purchases_summary = cur.fetchone()

        cur.execute(f"""
            SELECT customers.name, COUNT(*) as bills,
                   SUM(sales.amount) as total, SUM(sales.paid_amount) as paid,
                   SUM(sales.amount - sales.paid_amount) as pending
            FROM sales
            JOIN customers ON sales.customer_id = customers.id
            {sales_filter}
            GROUP BY customers.id ORDER BY total DESC
        """, (filter_val,))
        customer_breakdown = cur.fetchall()

        # All sales rows
        cur.execute(f"""
            SELECT sales.sale_date, customers.name as customer_name, sales.bill_no,
                   sales.amount, sales.paid_amount, (sales.amount - sales.paid_amount) as pending,
                   sales.notes
            FROM sales
            JOIN customers ON sales.customer_id = customers.id
            {sales_filter}
            ORDER BY sales.sale_date DESC
        """, (filter_val,))
        sales_rows = cur.fetchall()

        # All purchases rows
        cur.execute(f"""
            SELECT purchases.purchase_date, sellers.name as seller_name, purchases.bill_no,
                   purchases.payment_method, purchases.payment_ref,
                   purchases.product_description, purchases.amount, purchases.notes
            FROM purchases
            JOIN sellers ON purchases.seller_id = sellers.id
            {purchases_filter}
            ORDER BY purchases.purchase_date DESC
        """, (filter_val,))
        purchases_rows = cur.fetchall()
    finally:
        conn.close()

    total_sales     = float(sales_summary["total"] or 0)
    total_paid      = float(sales_summary["paid"]  or 0)
    total_pending   = total_sales - total_paid
    total_purchases = float(purchases_summary["total"] or 0)
    profit          = total_sales - total_purchases

    # ── Build workbook ──────────────────────────────────────
    wb = openpyxl.Workbook()

    # Styles
    hdr_font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill      = PatternFill("solid", fgColor="1E293B")
    sub_hdr_font  = Font(name="Calibri", bold=True, size=10)
    sub_hdr_fill  = PatternFill("solid", fgColor="E2E8F0")
    title_font    = Font(name="Calibri", bold=True, size=14)
    bold          = Font(name="Calibri", bold=True)
    normal        = Font(name="Calibri", size=10)
    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left",   vertical="center")
    thin          = Side(style="thin", color="CBD5E1")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    green_fill  = PatternFill("solid", fgColor="DCFCE7")
    amber_fill  = PatternFill("solid", fgColor="FEF3C7")
    red_fill    = PatternFill("solid", fgColor="FEE2E2")

    def style_header_row(ws, row_num, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = border

    def style_sub_header(ws, row_num, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = sub_hdr_font
            cell.fill      = sub_hdr_fill
            cell.alignment = center
            cell.border    = border

    def style_data_row(ws, row_num, cols, fill=None):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = normal
            cell.alignment = left
            cell.border    = border
            if fill:
                cell.fill = fill

    # ── Sheet 1: Summary ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20

    ws1["A1"] = f"Print Plus — Report ({period_label})"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:B1")
    ws1["A1"].alignment = center
    ws1.row_dimensions[1].height = 28

    ws1["A2"] = "Generated"
    ws1["B2"] = datetime.now().strftime("%d %b %Y, %I:%M %p")
    ws1["A2"].font = sub_hdr_font

    ws1.append([])

    headers = [["Metric", "Value"]]
    data = [
        ["Total Sales (₹)",         round(total_sales,     2)],
        ["Total Collected (₹)",     round(total_paid,      2)],
        ["Total Pending (₹)",       round(total_pending,   2)],
        ["Total Purchases (₹)",     round(total_purchases, 2)],
        ["Net Profit (₹)",          round(profit,          2)],
        ["No. of Sales Bills",      sales_summary["cnt"] or 0],
        ["No. of Purchases",        purchases_summary["cnt"] or 0],
    ]

    style_header_row(ws1, 4, 2)
    ws1["A4"] = "Metric"
    ws1["B4"] = "Value"

    for i, (label, val) in enumerate(data, start=5):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=val)
        fill = None
        if "Pending" in label:   fill = amber_fill
        if "Profit"  in label:   fill = green_fill if profit >= 0 else red_fill
        style_data_row(ws1, i, 2, fill)

    # ── Sheet 2: Customer Breakdown ─────────────────────────
    ws2 = wb.create_sheet("Customer Breakdown")
    for col, w in zip("ABCDEF", [30, 12, 18, 18, 18, 5]):
        ws2.column_dimensions[col].width = w

    ws2["A1"] = f"Customer Breakdown — {period_label}"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:F1")
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 24

    style_header_row(ws2, 2, 6)
    for col, hdr in enumerate(["Customer", "Bills", "Total (₹)", "Collected (₹)", "Pending (₹)", "#"], 1):
        ws2.cell(row=2, column=col, value=hdr)

    for i, c in enumerate(customer_breakdown, start=3):
        pending_val = float(c["pending"] or 0)
        fill = amber_fill if pending_val > 0.01 else green_fill
        ws2.cell(row=i, column=1, value=c["name"])
        ws2.cell(row=i, column=2, value=c["bills"])
        ws2.cell(row=i, column=3, value=round(float(c["total"] or 0), 2))
        ws2.cell(row=i, column=4, value=round(float(c["paid"]  or 0), 2))
        ws2.cell(row=i, column=5, value=round(pending_val, 2))
        ws2.cell(row=i, column=6, value=i - 2)
        style_data_row(ws2, i, 6, fill)

    # ── Sheet 3: All Sales ───────────────────────────────────
    ws3 = wb.create_sheet("Sales")
    for col, w in zip("ABCDEFG", [14, 28, 16, 14, 14, 14, 35]):
        ws3.column_dimensions[col].width = w

    ws3["A1"] = f"Sales — {period_label}"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:G1")
    ws3["A1"].alignment = center
    ws3.row_dimensions[1].height = 24

    style_header_row(ws3, 2, 7)
    for col, hdr in enumerate(["Date", "Customer", "Bill No", "Amount (₹)", "Paid (₹)", "Pending (₹)", "Notes"], 1):
        ws3.cell(row=2, column=col, value=hdr)

    for i, r in enumerate(sales_rows, start=3):
        pending_val = float(r["pending"] or 0)
        fill = amber_fill if pending_val > 0.01 else green_fill
        ws3.cell(row=i, column=1, value=r["sale_date"])
        ws3.cell(row=i, column=2, value=r["customer_name"])
        ws3.cell(row=i, column=3, value=r["bill_no"])
        ws3.cell(row=i, column=4, value=round(float(r["amount"] or 0), 2))
        ws3.cell(row=i, column=5, value=round(float(r["paid_amount"] or 0), 2))
        ws3.cell(row=i, column=6, value=round(pending_val, 2))
        ws3.cell(row=i, column=7, value=r["notes"] or "")
        style_data_row(ws3, i, 7, fill)

    # ── Sheet 4: All Purchases ───────────────────────────────
    ws4 = wb.create_sheet("Purchases")
    for col, w in zip("ABCDEFGH", [14, 28, 16, 14, 18, 30, 14, 30]):
        ws4.column_dimensions[col].width = w

    ws4["A1"] = f"Purchases — {period_label}"
    ws4["A1"].font = title_font
    ws4.merge_cells("A1:H1")
    ws4["A1"].alignment = center
    ws4.row_dimensions[1].height = 24

    style_header_row(ws4, 2, 8)
    for col, hdr in enumerate(["Date", "Seller", "Bill No", "Amount (₹)", "Payment", "Ref", "Description", "Notes"], 1):
        ws4.cell(row=2, column=col, value=hdr)

    for i, r in enumerate(purchases_rows, start=3):
        ws4.cell(row=i, column=1, value=r["purchase_date"])
        ws4.cell(row=i, column=2, value=r["seller_name"])
        ws4.cell(row=i, column=3, value=r["bill_no"] or "")
        ws4.cell(row=i, column=4, value=round(float(r["amount"] or 0), 2))
        ws4.cell(row=i, column=5, value=r["payment_method"] or "")
        ws4.cell(row=i, column=6, value=r["payment_ref"] or "")
        ws4.cell(row=i, column=7, value=r["product_description"] or "")
        ws4.cell(row=i, column=8, value=r["notes"] or "")
        style_data_row(ws4, i, 8)

    # ── Send file ────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"PrintPlus_Report_{period_label}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ----------------------------
# Database Backup
# ----------------------------
@app.route("/backup")
def backup_db():
    import shutil, os
    from database import get_base_path
    db_path = os.path.join(get_base_path(), "accounts.db")
    backup_dir = os.path.join(get_base_path(), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"accounts_backup_{timestamp}.db"
    backup_path = os.path.join(backup_dir, backup_name)
    shutil.copy2(db_path, backup_path)
    # Keep only last 10 backups
    all_backups = sorted([f for f in os.listdir(backup_dir) if f.endswith(".db")])
    while len(all_backups) > 10:
        os.remove(os.path.join(backup_dir, all_backups.pop(0)))
    flash(f"Backup saved: {backup_name} ✅", "success")
    return redirect(url_for("home"))

if __name__ == "__main__":
    import threading
    import webbrowser
    import time

    def open_browser():
        time.sleep(2)
        webbrowser.open_new("http://127.0.0.1:5000/")
        # FIXED: removed broken db.create_all() call that was here — init_db() at the top handles this

    threading.Thread(target=open_browser).start()

    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)
